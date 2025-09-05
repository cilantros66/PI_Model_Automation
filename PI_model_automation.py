# TOOL 1 - GeneratePi

# Version 1: Pi functionality added.
# Version 2: -Corrected discharge import to use '"average","time-weighted"'
#            -Added option to make new dfs0 (without extending existing).
#            -Fill with delete value, not 0 and do not do addition/multiplication on delete value
#            -Note new rainfall will give an error because rainfall depth + step accumulated fails but that is what it should be.


import openpyxl
from openpyxl import Workbook
from datetime import datetime
from datetime import timedelta
import os
import pandas as pd
import numpy as np
import xlwings as xw
import mikeio
from mikecore.DfsFile import DataValueType
from mikeio import ItemInfo, EUMType, EUMUnit
from PI_model_automation_variables import * #import everything

df_input_all = pd.read_excel(inputsheet_path)

delete_value = -1e-35

if generate_pi:
    timestep_dict = {}
    timestep_dict["5m"]=300
    
    if custom_end_date == '':
        pi_end_time = datetime.today().replace(minute=0, second=0, microsecond=0) # today's time round down to nearest hour
    else:
        pi_end_time = custom_end_date
    interval = "5m"
    
    pi_dict = {}
    pi_dict['HGL'] = '"average","time-weighted"'
    pi_dict['Rainfall'] = '"total","event-weighted"'
    pi_dict['Flow'] = '"average","time-weighted"'
    
    dfs_dict = {}
    dfs_dict['HGL'] = [EUMType.Water_Level,DataValueType.Instantaneous]
    dfs_dict['Rainfall'] = [EUMType.Rainfall_Depth,DataValueType.StepAccumulated]#This gives an error but better than wrong value. To be reported.
    dfs_dict['Flow'] = [EUMType.Discharge,DataValueType.Instantaneous]
    
    for index, row in df_input_all.iterrows():
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
    
        name = row["Name"]
        tag = row["Tag"]
        item_type = row["Type"]
        unit = row["Unit"]
        desc = f"{type} ({unit}) in gauge {name}"

        if extend_existing:    
            dfs0_path = row["DFS0 Path"]
            dfs = mikeio.read(dfs0_path)
            dfs0_end_time = dfs.end_time
            pi_start_time = dfs0_end_time + timedelta(minutes=5)
        else:
            pi_start_time = datetime.strptime(custom_start_date, '%Y-%m-%d')
            
        
        ws["A1"] = "Start"
        ws["B1"] = pi_start_time.strftime("%Y-%m-%d %H:%M:%S")
        ws["A2"] = "End"
        ws["B2"] = pi_end_time.strftime("%Y-%m-%d %H:%M:%S")
    
        # Duration
        duration =  pi_end_time -  pi_start_time
        duration_seconds = duration.total_seconds()
        timestep_seconds = timestep_dict[interval]
        timestep_number = duration_seconds/timestep_seconds - 1
        last_row = int(timestep_number + 10)
    
        # Interval
        ws["A3"] = "Interval"
        #interval = input("Enter Interval - e.g.5m : ")
        ws["B3"] = interval
    
        # Name (used for Excel file name)
        ws["A5"] = "Name"
        # name = input("Enter Name: ")
        ws["B5"] = name
    
        # Tag
        ws["A6"] = "Tag"
        ws["B6"] = tag
    
        # Desc
        ws["A7"] = "Desc"
        ws["B7"] = desc
    
        # Unit
        ws["A8"] = "Unit"
        ws["B8"] = unit
    
        # Array formula
        ws["A10"] = r'=PIAdvCalcDat(Sheet1!$B$6,Sheet1!$B$1,Sheet1!$B$2,Sheet1!$B$3,' + pi_dict[item_type] + ',0,1,65,"\\gvprdhist01")'
        ws.formula_attributes["A10"] = {'t': 'array', 'ref': f"A10:B{last_row}"} 
        for i in range(10,last_row + 1):
            ws[f'A{i}'].number_format = 'yyyy-mm-dd hh:mm:ss'
        ws.column_dimensions['A'].width = 25
        # Freeze the first column
        ws.freeze_panes = ws['A10']
    
        # Save the file
        file_path = f"{output_folder}\\{name}-{item_type}.xlsx"
        wb.save(file_path)
                    
        # Open, save, and close the workbook
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        wb.save()
        wb.close()
        app.quit()
    
    #     os.startfile(file_name)  
    #     Ctrl + / the above line for debugging only
    
        print(f"Excel file '{file_path}' has been created!!")
        
# Tool 2: Generate dfs0 files
        
if generate_dfs0:
    for dfs0_path in list(df_input_all["DFS0 Path"].unique()):
        
        #Should be moved out of this loop so multiple PI sheets can go into 1 dfs0
        if extend_existing:
            df_input = df_input_all[df_input_all["DFS0 Path"]==dfs0_path]
            dfs = mikeio.read(dfs0_path)
            dfs_df = dfs.to_dataframe()
        else:
            df_input = df_input_all
        
        for index, row in df_input.iterrows():
            if not pd.isna(row.Tag):
                name = row["Name"]
                item1_name = row["DFS0 Item 1 Name"]
                var_type = row["Type"]
                pi_file_path = output_folder + "\\" + name + "-" + var_type + ".xlsx"
                print(pi_file_path)
    
                # read in input pi spreadsheet
                pi_df = pd.read_excel(pi_file_path, skiprows = 4)
    
    
                unit = pi_df.iloc[2, 1]
                pi_df.drop(pi_df.index[:4], inplace = True)
                pi_df.reset_index(drop = True, inplace = True)
                pi_df.rename(columns = {"Name":"DateTimeInitial"}, inplace = True)
                pi_df['DateTime'] = pd.to_datetime(pi_df.DateTimeInitial, errors='coerce')
                pi_df[name] = pd.to_numeric(pi_df[name], errors='coerce').fillna(delete_value)
                pi_df.dropna(inplace=True)
                pi_df.drop(columns=['DateTimeInitial'],inplace=True)
                
                if not pd.isna(row['DFS0 Item 1 Addition']):
                    pi_df[name] = np.where(pi_df[name] != delete_value, pi_df[name] + row['DFS0 Item 1 Addition'], pi_df[name])
                
                if not pd.isna(row['DFS0 Item 1 Multiplier']):
                    pi_df[name] = np.where(pi_df[name] != delete_value, pi_df[name] * row['DFS0 Item 1 Multiplier'], pi_df[name])
                    
                if extend_existing:
                    pi_df.rename(columns={name:name + '_Transfer'},inplace=True)
                      
                    pi_end_time = pi_df.DateTime.max()
                    rng = pd.date_range(dfs_df.index.min(),pi_end_time,freq='300s')
        
                    #join df with dfs_df by DateTime, then transfer the VW14 values from df to dfs_df. 
            
                    #Make sure it is an outer join to maintain all dfs0 times in the dfs_df
                
                    ix = pd.DatetimeIndex(rng)
                    dfs_df = dfs_df.reindex(ix)
                    dfs_df['DateTime'] = dfs_df.index        
                    dfs_df = pd.merge(dfs_df,pi_df,how='left',on=['DateTime'])        
                    dfs_df.set_index('DateTime',inplace=True)        
                    dfs_df[item1_name] = dfs_df[item1_name].fillna(dfs_df[name + '_Transfer'])       
                    dfs_df.drop(columns=[name + '_Transfer'],inplace=True)                    
                    dfs_df.to_dfs0(f"{output_folder}\\{name}_Extended.dfs0", 
                                   items=dfs.items, title=name)
                else:
                    eum_type = dfs_dict[var_type][0]
                    data_value_type = dfs_dict[var_type][1]
                    # item = ItemInfo(f'{name} - {var_type}',eum_type,data_value_type)
                    item = ItemInfo(f'{name} - {var_type}',eum_type)
                    items = [item]
                    pi_df.set_index('DateTime',inplace=True)
                    pi_df.to_dfs0(f"{output_folder}\\{name} - {var_type}.dfs0",
                                  items=[item],title=name)
                    

    